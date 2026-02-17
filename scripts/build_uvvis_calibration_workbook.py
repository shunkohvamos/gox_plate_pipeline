#!/usr/bin/env python3
"""Build an Excel workbook for GOx UV-Vis calibration and quick concentration calculation."""

from __future__ import annotations

import argparse
import csv
import hashlib
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


@dataclass(frozen=True)
class SpectrumPoint:
    wavelength_nm: float
    absorbance: float


@dataclass(frozen=True)
class CalibrationRecord:
    concentration_mg_ml: float
    source_file: str
    wavelength_low_nm: float
    absorbance_low: float
    wavelength_high_nm: float
    absorbance_high: float
    absorbance_target: float
    use_for_fit: int


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create a UV-Vis calibration Excel workbook with formulas and a calibration chart."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("research strategy/experimental method/UV-VIS検量線用データ"),
        help="Directory containing UV-Vis CSV files (default: %(default)s)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path (default: <input-dir>/UVVIS_GOx_calibration_<run_id>.xlsx)",
    )
    parser.add_argument(
        "--target-wavelength",
        type=float,
        default=277.0,
        help="Target wavelength in nm for calibration extraction (default: %(default)s)",
    )
    return parser.parse_args()


def parse_concentration_from_filename(path: Path) -> float:
    """Parse concentration from filenames like Absorbance_0_0_5.csv -> 0.5 mg/mL."""
    stem = path.stem
    if not stem.startswith("Absorbance_"):
        raise ValueError(f"Unexpected file name format: {path.name}")
    token = stem[len("Absorbance_") :]
    # The first "0_" is device/export prefix in this dataset.
    if token.startswith("0_"):
        token = token[2:]
    parts = token.split("_")
    if not parts or any(not p.isdigit() for p in parts):
        raise ValueError(f"Cannot parse concentration from file name: {path.name}")
    if len(parts) == 1:
        return float(parts[0])
    return float(f"{parts[0]}.{''.join(parts[1:])}")


def load_spectrum(path: Path) -> list[SpectrumPoint]:
    points: list[SpectrumPoint] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) < 2:
                continue
            try:
                wavelength_nm = float(row[0].strip())
                absorbance = float(row[1].strip())
            except ValueError:
                continue
            points.append(SpectrumPoint(wavelength_nm=wavelength_nm, absorbance=absorbance))
    if len(points) < 2:
        raise ValueError(f"No spectral points were parsed from {path}")
    return points


def interpolate_at_wavelength(points: Iterable[SpectrumPoint], target_wavelength: float) -> tuple[float, float, float, float, float]:
    ordered = sorted(points, key=lambda p: p.wavelength_nm)
    if target_wavelength < ordered[0].wavelength_nm or target_wavelength > ordered[-1].wavelength_nm:
        raise ValueError(
            f"Target wavelength {target_wavelength} nm is outside range "
            f"{ordered[0].wavelength_nm}..{ordered[-1].wavelength_nm} nm"
        )
    for i in range(len(ordered) - 1):
        left = ordered[i]
        right = ordered[i + 1]
        if left.wavelength_nm <= target_wavelength <= right.wavelength_nm:
            if right.wavelength_nm == left.wavelength_nm:
                return (
                    left.wavelength_nm,
                    left.absorbance,
                    right.wavelength_nm,
                    right.absorbance,
                    left.absorbance,
                )
            ratio = (target_wavelength - left.wavelength_nm) / (right.wavelength_nm - left.wavelength_nm)
            absorbance_target = left.absorbance + ratio * (right.absorbance - left.absorbance)
            return (
                left.wavelength_nm,
                left.absorbance,
                right.wavelength_nm,
                right.absorbance,
                absorbance_target,
            )
    raise ValueError(f"Could not find interpolation bracket around {target_wavelength} nm")


def infer_default_use_flags(records_sorted: list[CalibrationRecord]) -> list[int]:
    """Mark possible saturation point at highest concentration as excluded by default."""
    flags = [1] * len(records_sorted)
    if len(records_sorted) < 3:
        return flags
    deltas = [
        records_sorted[i + 1].absorbance_target - records_sorted[i].absorbance_target
        for i in range(len(records_sorted) - 1)
    ]
    previous_positive = [d for d in deltas[:-1] if d > 0]
    if not previous_positive:
        return flags
    mean_previous = sum(previous_positive) / len(previous_positive)
    if mean_previous > 0 and deltas[-1] <= 0.25 * mean_previous:
        flags[-1] = 0
    return flags


def sha256sum(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def autosize_columns(ws, start_col: int, end_col: int) -> None:
    for col_idx in range(start_col, end_col + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 45)


def build_workbook(
    records: list[CalibrationRecord],
    raw_spectra: dict[str, list[SpectrumPoint]],
    source_files: list[Path],
    target_wavelength: float,
    run_id: str,
) -> Workbook:
    wb = Workbook()
    ws_cal = wb.active
    ws_cal.title = "Calibration"

    title_font = Font(name="Arial", size=12, bold=True)
    header_font = Font(name="Arial", size=10, bold=True)
    body_font = Font(name="Arial", size=10)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    input_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    center = Alignment(horizontal="center", vertical="center")

    ws_cal["A1"] = "GOx UV-Vis Calibration (Abs at target wavelength)"
    ws_cal["A1"].font = title_font
    ws_cal["A2"] = "run_id"
    ws_cal["B2"] = run_id
    ws_cal["A3"] = "target_wavelength_nm"
    ws_cal["B3"] = target_wavelength

    headers = [
        "Concentration_mg_mL",
        "Source_File",
        "Lambda_Low_nm",
        "Abs_Low",
        "Lambda_High_nm",
        "Abs_High",
        "Abs_Target",
        "Use_For_Fit (1/0)",
        "Fit_Abs",
        "Residual (Obs-Fit)",
    ]
    start_row = 5
    for idx, header in enumerate(headers, start=1):
        cell = ws_cal.cell(row=start_row, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for i, record in enumerate(records, start=start_row + 1):
        ws_cal.cell(row=i, column=1, value=record.concentration_mg_ml).font = body_font
        ws_cal.cell(row=i, column=2, value=record.source_file).font = body_font
        ws_cal.cell(row=i, column=3, value=record.wavelength_low_nm).font = body_font
        ws_cal.cell(row=i, column=4, value=record.absorbance_low).font = body_font
        ws_cal.cell(row=i, column=5, value=record.wavelength_high_nm).font = body_font
        ws_cal.cell(row=i, column=6, value=record.absorbance_high).font = body_font
        ws_cal.cell(row=i, column=7, value=record.absorbance_target).font = body_font
        fit_flag_cell = ws_cal.cell(row=i, column=8, value=record.use_for_fit)
        fit_flag_cell.font = body_font
        fit_flag_cell.fill = input_fill
        ws_cal.cell(row=i, column=9, value=f"=IF(H{i}=1,$M$11*A{i}+$M$12,NA())").font = body_font
        ws_cal.cell(row=i, column=10, value=f"=IF(H{i}=1,G{i}-I{i},NA())").font = body_font

    end_row = start_row + len(records)
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
    ws_cal.add_data_validation(dv)
    dv.add(f"H{start_row + 1}:H{end_row}")

    ws_cal["L4"] = "Regression Stats (selected points only)"
    ws_cal["L4"].font = header_font
    stats = [
        ("N_used", f"=SUM(H{start_row + 1}:H{end_row})"),
        ("Sum_x", f"=SUMPRODUCT(A{start_row + 1}:A{end_row},H{start_row + 1}:H{end_row})"),
        ("Sum_y", f"=SUMPRODUCT(G{start_row + 1}:G{end_row},H{start_row + 1}:H{end_row})"),
        ("Sum_xx", f"=SUMPRODUCT(A{start_row + 1}:A{end_row},A{start_row + 1}:A{end_row},H{start_row + 1}:H{end_row})"),
        ("Sum_yy", f"=SUMPRODUCT(G{start_row + 1}:G{end_row},G{start_row + 1}:G{end_row},H{start_row + 1}:H{end_row})"),
        ("Sum_xy", f"=SUMPRODUCT(A{start_row + 1}:A{end_row},G{start_row + 1}:G{end_row},H{start_row + 1}:H{end_row})"),
        (
            "Slope_abs_per_mg",
            "=IF(M5<2,NA(),(M10-(M6*M7)/M5)/(M8-(M6^2)/M5))",
        ),
        ("Intercept_abs", "=IF(M5<2,NA(),(M7-M11*M6)/M5)"),
        (
            "R_squared",
            "=IF(OR(M5<2,(M5*M8-M6^2)=0,(M5*M9-M7^2)=0),NA(),((M5*M10-M6*M7)^2)/((M5*M8-M6^2)*(M5*M9-M7^2)))",
        ),
        ("Slope_mg_per_abs", "=IFERROR(1/M11,NA())"),
        ("Intercept_mg", "=IFERROR(-M12/M11,NA())"),
    ]
    for offset, (label, formula) in enumerate(stats, start=5):
        ws_cal[f"L{offset}"] = label
        ws_cal[f"L{offset}"].font = header_font
        ws_cal[f"M{offset}"] = formula
        ws_cal[f"M{offset}"].font = body_font

    ws_cal["P5"] = "FitLine_X"
    ws_cal["Q5"] = "FitLine_Y"
    ws_cal["P6"] = f"=MIN(A{start_row + 1}:A{end_row})"
    ws_cal["P7"] = f"=MAX(A{start_row + 1}:A{end_row})"
    ws_cal["Q6"] = "=IFERROR($M$11*P6+$M$12,NA())"
    ws_cal["Q7"] = "=IFERROR($M$11*P7+$M$12,NA())"

    chart = ScatterChart()
    chart.title = "Calibration Curve at target wavelength"
    chart.x_axis.title = "GOx concentration (mg/mL)"
    chart.y_axis.title = "Absorbance"
    chart.style = 13
    chart.height = 10
    chart.width = 15

    xvalues = Reference(ws_cal, min_col=1, min_row=start_row + 1, max_row=end_row)
    yvalues = Reference(ws_cal, min_col=7, min_row=start_row + 1, max_row=end_row)
    obs_series = Series(yvalues, xvalues, title="Observed Abs (all points)")
    obs_series.marker.symbol = "circle"
    obs_series.marker.size = 8
    obs_series.graphicalProperties.line.noFill = True
    chart.series.append(obs_series)

    fit_x = Reference(ws_cal, min_col=16, min_row=6, max_row=7)
    fit_y = Reference(ws_cal, min_col=17, min_row=6, max_row=7)
    fit_series = Series(fit_y, fit_x, title="Linear fit (Use_For_Fit=1)")
    fit_series.graphicalProperties.line.solidFill = "1F4E78"
    fit_series.graphicalProperties.line.width = 25000
    fit_series.marker = None
    chart.series.append(fit_series)
    ws_cal.add_chart(chart, "A12")

    ws_cal["A29"] = "Note"
    ws_cal["A30"] = "Edit 'Use_For_Fit' (1/0) to change fitting range. QuickCalc uses the selected-fit coefficients."
    ws_cal["A31"] = "A possible saturation point at the highest concentration is auto-excluded when its absorbance increment is very small."
    ws_cal["A29"].font = header_font
    ws_cal["A30"].font = body_font
    ws_cal["A31"].font = body_font

    autosize_columns(ws_cal, 1, 13)
    ws_cal.freeze_panes = "A6"

    ws_quick = wb.create_sheet(title="QuickCalc")
    ws_quick["A1"] = "GOx Concentration Quick Calculator"
    ws_quick["A1"].font = title_font
    rows = [
        ("Input: Absorbance at target wavelength", None),
        ("Measured Abs", None),
        ("Estimated GOx concentration (mg/mL)", "=IF(B3=\"\",\"\",(B3-Calibration!$M$12)/Calibration!$M$11)"),
        ("Target concentration (mg/mL)", 1.0),
        ("Dilution factor (stock/target)", "=IF(OR(B4=\"\",B5=\"\",B5=0),\"\",B4/B5)"),
        ("Desired final volume (uL)", 1000),
        ("Stock volume to pipette (uL)", "=IF(OR(B6=\"\",B6=0,B7=\"\"),\"\",B7/B6)"),
        ("Buffer volume to add (uL)", "=IF(OR(B7=\"\",B8=\"\"),\"\",B7-B8)"),
        ("", None),
        ("Optional absorbance-normalization target", None),
        ("Target Abs", 0.3),
        ("Volume correction factor (measured/target Abs)", "=IF(OR(B3=\"\",B12=\"\",B12=0),\"\",B3/B12)"),
    ]
    start = 2
    for idx, (label, value) in enumerate(rows, start=start):
        ws_quick.cell(row=idx, column=1, value=label).font = body_font
        if value is not None:
            ws_quick.cell(row=idx, column=2, value=value).font = body_font

    ws_quick["B3"].fill = input_fill
    ws_quick["B5"].fill = input_fill
    ws_quick["B7"].fill = input_fill
    ws_quick["B12"].fill = input_fill
    ws_quick["A15"] = "Current fit coefficients (from Calibration)"
    ws_quick["A15"].font = header_font
    ws_quick["A16"] = "Abs = slope * Conc + intercept"
    ws_quick["A17"] = "slope_abs_per_mg"
    ws_quick["B17"] = "=Calibration!$M$11"
    ws_quick["A18"] = "intercept_abs"
    ws_quick["B18"] = "=Calibration!$M$12"
    ws_quick["A19"] = "R_squared"
    ws_quick["B19"] = "=Calibration!$M$13"
    autosize_columns(ws_quick, 1, 2)

    ws_raw = wb.create_sheet(title="RawSpectra")
    ws_raw["A1"] = "Wavelength_nm"
    ws_raw["A1"].font = header_font
    sorted_keys = sorted(raw_spectra.keys(), key=lambda x: float(x))
    for col_idx, key in enumerate(sorted_keys, start=2):
        ws_raw.cell(row=1, column=col_idx, value=f"Abs_{key}_mg_mL").font = header_font

    base_points = raw_spectra[sorted_keys[0]]
    for row_idx, point in enumerate(base_points, start=2):
        ws_raw.cell(row=row_idx, column=1, value=point.wavelength_nm).font = body_font
        for col_idx, key in enumerate(sorted_keys, start=2):
            points = raw_spectra[key]
            ws_raw.cell(row=row_idx, column=col_idx, value=points[row_idx - 2].absorbance).font = body_font
    autosize_columns(ws_raw, 1, len(sorted_keys) + 1)
    ws_raw.freeze_panes = "A2"

    ws_prov = wb.create_sheet(title="Provenance")
    ws_prov["A1"] = "run_id"
    ws_prov["B1"] = run_id
    ws_prov["A2"] = "generated_at"
    ws_prov["B2"] = datetime.now().isoformat(timespec="seconds")
    ws_prov["A3"] = "target_wavelength_nm"
    ws_prov["B3"] = target_wavelength
    ws_prov["A4"] = "extraction_method"
    ws_prov["B4"] = "Linear interpolation between nearest bracketing wavelengths"
    ws_prov["A6"] = "source_file"
    ws_prov["B6"] = "concentration_mg_mL"
    ws_prov["C6"] = "sha256"
    ws_prov["D6"] = "size_bytes"
    ws_prov["E6"] = "mtime_iso"
    for cell in ("A6", "B6", "C6", "D6", "E6"):
        ws_prov[cell].font = header_font
        ws_prov[cell].fill = header_fill
        ws_prov[cell].alignment = center
    for idx, file_path in enumerate(source_files, start=7):
        ws_prov.cell(row=idx, column=1, value=str(file_path))
        ws_prov.cell(row=idx, column=2, value=parse_concentration_from_filename(file_path))
        ws_prov.cell(row=idx, column=3, value=sha256sum(file_path))
        ws_prov.cell(row=idx, column=4, value=file_path.stat().st_size)
        ws_prov.cell(row=idx, column=5, value=datetime.fromtimestamp(file_path.stat().st_mtime).isoformat(timespec="seconds"))
    autosize_columns(ws_prov, 1, 5)

    return wb


def main() -> None:
    args = parse_args()
    input_dir: Path = args.input_dir
    target_wavelength: float = args.target_wavelength

    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory not found: {input_dir}")

    source_files = sorted(input_dir.glob("Absorbance_*.csv"))
    if not source_files:
        raise FileNotFoundError(f"No CSV files found under {input_dir}")

    temp_records: list[CalibrationRecord] = []
    raw_spectra: dict[str, list[SpectrumPoint]] = {}
    for file_path in source_files:
        concentration = parse_concentration_from_filename(file_path)
        points = load_spectrum(file_path)
        low_w, low_a, high_w, high_a, target_a = interpolate_at_wavelength(points, target_wavelength)
        temp_records.append(
            CalibrationRecord(
                concentration_mg_ml=concentration,
                source_file=file_path.name,
                wavelength_low_nm=low_w,
                absorbance_low=low_a,
                wavelength_high_nm=high_w,
                absorbance_high=high_a,
                absorbance_target=target_a,
                use_for_fit=1,
            )
        )
        raw_spectra[str(concentration)] = points

    temp_records.sort(key=lambda r: r.concentration_mg_ml)
    use_flags = infer_default_use_flags(temp_records)
    records = [
        CalibrationRecord(
            concentration_mg_ml=r.concentration_mg_ml,
            source_file=r.source_file,
            wavelength_low_nm=r.wavelength_low_nm,
            absorbance_low=r.absorbance_low,
            wavelength_high_nm=r.wavelength_high_nm,
            absorbance_high=r.absorbance_high,
            absorbance_target=r.absorbance_target,
            use_for_fit=use_flags[i],
        )
        for i, r in enumerate(temp_records)
    ]

    run_id = datetime.now().strftime("uvvis_cal_%Y%m%d_%H%M%S")
    output_path = (
        args.output
        if args.output is not None
        else input_dir / f"UVVIS_GOx_calibration_{run_id}.xlsx"
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook(
        records=records,
        raw_spectra=raw_spectra,
        source_files=source_files,
        target_wavelength=target_wavelength,
        run_id=run_id,
    )
    wb.save(output_path)
    print(f"Created workbook: {output_path}")
    print(f"run_id: {run_id}")
    print("Calibration points:")
    for r in records:
        print(
            f"  conc={r.concentration_mg_ml:.4g} mg/mL, "
            f"Abs@{target_wavelength:.2f}nm={r.absorbance_target:.6f}, use_for_fit={r.use_for_fit}"
        )


if __name__ == "__main__":
    main()
