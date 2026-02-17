#!/usr/bin/env python3
"""Build calibration sheets from a user-prepared UV-Vis summary Excel file."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


@dataclass(frozen=True)
class DataPoint:
    concentration_mg_ml: float
    absorbance: float


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create UV-Vis calibration curve sheets from a summary Excel workbook."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("research strategy/experimental method/UV-VIS検量線用データ/検量線作成.xlsx"),
        help="Input workbook path that already contains concentration and absorbance table.",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Sheet name containing summary table (default: first sheet).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output workbook path (default: <input_stem>_with_calibration.xlsx).",
    )
    parser.add_argument(
        "--target-wavelength",
        type=float,
        default=278.64,
        help="Target wavelength used to get absorbance values (nm).",
    )
    return parser.parse_args()


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def extract_points(ws) -> list[DataPoint]:
    header_row = None
    conc_col = None
    abs_col = None
    max_scan_row = min(ws.max_row, 200)
    max_scan_col = min(ws.max_column, 40)

    for r in range(1, max_scan_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_scan_col + 1)]
        for c, value in enumerate(row_vals, start=1):
            text = str(value).lower() if value is not None else ""
            if "concentration" in text:
                conc_col = c
            if "absorbance" in text:
                abs_col = c
        if conc_col is not None and abs_col is not None:
            header_row = r
            break

    if header_row is None:
        raise ValueError("Could not find header row with both 'Concentration' and 'Absorbance'.")

    points: list[DataPoint] = []
    for r in range(header_row + 1, ws.max_row + 1):
        conc = _to_float(ws.cell(r, conc_col).value)
        absorbance = _to_float(ws.cell(r, abs_col).value)
        if conc is None and absorbance is None:
            continue
        if conc is None or absorbance is None:
            continue
        points.append(DataPoint(concentration_mg_ml=conc, absorbance=absorbance))

    if len(points) < 2:
        raise ValueError("Need at least 2 data points to build a calibration curve.")

    points.sort(key=lambda p: p.concentration_mg_ml)
    return points


def infer_use_flags(points: list[DataPoint]) -> list[int]:
    flags = [1] * len(points)
    if len(points) < 3:
        return flags
    deltas = [points[i + 1].absorbance - points[i].absorbance for i in range(len(points) - 1)]
    previous_positive = [d for d in deltas[:-1] if d > 0]
    if not previous_positive:
        return flags
    mean_previous = sum(previous_positive) / len(previous_positive)
    if mean_previous > 0 and deltas[-1] <= 0.25 * mean_previous:
        flags[-1] = 0
    return flags


def autosize_columns(ws, start_col: int, end_col: int) -> None:
    for col_idx in range(start_col, end_col + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 42)


def remove_sheet_if_exists(wb, name: str) -> None:
    if name in wb.sheetnames:
        ws = wb[name]
        wb.remove(ws)


def add_calibration_sheet(wb, points: list[DataPoint], use_flags: list[int], target_wavelength: float) -> None:
    remove_sheet_if_exists(wb, "CalibrationCurve")
    ws = wb.create_sheet("CalibrationCurve")

    title_font = Font(name="Arial", size=12, bold=True)
    header_font = Font(name="Arial", size=10, bold=True)
    body_font = Font(name="Arial", size=10)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    input_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    center = Alignment(horizontal="center", vertical="center")

    run_id = datetime.now().strftime("uvvis_xlsx_%Y%m%d_%H%M%S")

    ws["A1"] = "GOx UV-Vis Calibration Curve"
    ws["A1"].font = title_font
    ws["A2"] = "run_id"
    ws["B2"] = run_id
    ws["A3"] = "target_wavelength_nm"
    ws["B3"] = target_wavelength

    headers = ["Concentration_mg_mL", "Absorbance", "Use_For_Fit (1/0)", "Fit_Abs", "Residual (Obs-Fit)"]
    start_row = 5
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for i, point in enumerate(points, start=start_row + 1):
        ws.cell(row=i, column=1, value=point.concentration_mg_ml).font = body_font
        ws.cell(row=i, column=2, value=point.absorbance).font = body_font
        cell_flag = ws.cell(row=i, column=3, value=use_flags[i - (start_row + 1)])
        cell_flag.font = body_font
        cell_flag.fill = input_fill
        ws.cell(row=i, column=4, value=f"=IF(C{i}=1,$L$11*A{i}+$L$12,NA())").font = body_font
        ws.cell(row=i, column=5, value=f"=IF(C{i}=1,B{i}-D{i},NA())").font = body_font

    end_row = start_row + len(points)
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"C{start_row + 1}:C{end_row}")

    ws["K4"] = "Regression Stats (Use_For_Fit=1)"
    ws["K4"].font = header_font
    stats = [
        ("N_used", f"=SUM(C{start_row + 1}:C{end_row})"),
        ("Sum_x", f"=SUMPRODUCT(A{start_row + 1}:A{end_row},C{start_row + 1}:C{end_row})"),
        ("Sum_y", f"=SUMPRODUCT(B{start_row + 1}:B{end_row},C{start_row + 1}:C{end_row})"),
        ("Sum_xx", f"=SUMPRODUCT(A{start_row + 1}:A{end_row},A{start_row + 1}:A{end_row},C{start_row + 1}:C{end_row})"),
        ("Sum_yy", f"=SUMPRODUCT(B{start_row + 1}:B{end_row},B{start_row + 1}:B{end_row},C{start_row + 1}:C{end_row})"),
        ("Sum_xy", f"=SUMPRODUCT(A{start_row + 1}:A{end_row},B{start_row + 1}:B{end_row},C{start_row + 1}:C{end_row})"),
        ("Slope_abs_per_mg", "=IF(L5<2,NA(),(L10-(L6*L7)/L5)/(L8-(L6^2)/L5))"),
        ("Intercept_abs", "=IF(L5<2,NA(),(L7-L11*L6)/L5)"),
        (
            "R_squared",
            "=IF(OR(L5<2,(L5*L8-L6^2)=0,(L5*L9-L7^2)=0),NA(),((L5*L10-L6*L7)^2)/((L5*L8-L6^2)*(L5*L9-L7^2)))",
        ),
        ("Slope_mg_per_abs", "=IFERROR(1/L11,NA())"),
        ("Intercept_mg", "=IFERROR(-L12/L11,NA())"),
    ]
    for idx, (name, formula) in enumerate(stats, start=5):
        ws[f"K{idx}"] = name
        ws[f"K{idx}"].font = header_font
        ws[f"L{idx}"] = formula
        ws[f"L{idx}"].font = body_font

    ws["N5"] = "FitLine_X"
    ws["O5"] = "FitLine_Y"
    ws["N6"] = f"=MIN(A{start_row + 1}:A{end_row})"
    ws["N7"] = f"=MAX(A{start_row + 1}:A{end_row})"
    ws["O6"] = "=IFERROR($L$11*N6+$L$12,NA())"
    ws["O7"] = "=IFERROR($L$11*N7+$L$12,NA())"

    chart = ScatterChart()
    chart.title = "Calibration Curve"
    chart.x_axis.title = "GOx concentration (mg/mL)"
    chart.y_axis.title = "Absorbance"
    chart.style = 13
    chart.height = 10
    chart.width = 15

    obs_x = Reference(ws, min_col=1, min_row=start_row + 1, max_row=end_row)
    obs_y = Reference(ws, min_col=2, min_row=start_row + 1, max_row=end_row)
    obs_series = Series(obs_y, obs_x, title="Observed")
    obs_series.marker.symbol = "circle"
    obs_series.marker.size = 8
    obs_series.graphicalProperties.line.noFill = True
    chart.series.append(obs_series)

    fit_x = Reference(ws, min_col=14, min_row=6, max_row=7)
    fit_y = Reference(ws, min_col=15, min_row=6, max_row=7)
    fit_series = Series(fit_y, fit_x, title="Linear fit")
    fit_series.marker = None
    fit_series.graphicalProperties.line.solidFill = "1F4E78"
    fit_series.graphicalProperties.line.width = 25000
    chart.series.append(fit_series)

    ws.add_chart(chart, "A11")
    ws["A28"] = "Note"
    ws["A29"] = "You can switch each point on/off via Use_For_Fit and the fit/QuickCalc update automatically."
    ws["A30"] = "Highest concentration may be excluded by default when saturation-like behavior is detected."
    ws["A28"].font = header_font
    ws["A29"].font = body_font
    ws["A30"].font = body_font

    ws.freeze_panes = "A6"
    autosize_columns(ws, 1, 12)


def add_quickcalc_sheet(wb) -> None:
    remove_sheet_if_exists(wb, "QuickCalc")
    ws = wb.create_sheet("QuickCalc")

    title_font = Font(name="Arial", size=12, bold=True)
    header_font = Font(name="Arial", size=10, bold=True)
    body_font = Font(name="Arial", size=10)
    input_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    ws["A1"] = "GOx Concentration Quick Calculator"
    ws["A1"].font = title_font

    rows = [
        ("Input: measured absorbance", None),
        ("Measured Abs", None),
        ("Estimated GOx concentration (mg/mL)", "=IF(B3=\"\",\"\",(B3-CalibrationCurve!$L$12)/CalibrationCurve!$L$11)"),
        ("Target concentration (mg/mL)", 1.0),
        ("Dilution factor (stock/target)", "=IF(OR(B4=\"\",B5=\"\",B5=0),\"\",B4/B5)"),
        ("Desired final volume (uL)", 1000),
        ("Stock volume to pipette (uL)", "=IF(OR(B6=\"\",B6=0,B7=\"\"),\"\",B7/B6)"),
        ("Buffer volume to add (uL)", "=IF(OR(B7=\"\",B8=\"\"),\"\",B7-B8)"),
        ("", None),
        ("Optional absorbance-normalization mode", None),
        ("Target Abs", 0.3),
        ("Volume correction factor (measured/target)", "=IF(OR(B3=\"\",B12=\"\",B12=0),\"\",B3/B12)"),
        ("", None),
        ("Current fit summary", None),
        ("Slope_abs_per_mg", "=CalibrationCurve!$L$11"),
        ("Intercept_abs", "=CalibrationCurve!$L$12"),
        ("R_squared", "=CalibrationCurve!$L$13"),
    ]
    for r, (label, value) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=label).font = body_font
        if value is not None:
            ws.cell(row=r, column=2, value=value).font = body_font
    ws["A15"].font = header_font

    for addr in ("B3", "B5", "B7", "B12"):
        ws[addr].fill = input_fill

    autosize_columns(ws, 1, 2)


def main() -> None:
    args = parse_args()
    input_path = args.input
    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    output_path = (
        args.output
        if args.output is not None
        else input_path.with_name(f"{input_path.stem}_with_calibration.xlsx")
    )

    wb = load_workbook(input_path)
    source_sheet = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    points = extract_points(source_sheet)
    use_flags = infer_use_flags(points)

    add_calibration_sheet(wb, points=points, use_flags=use_flags, target_wavelength=args.target_wavelength)
    add_quickcalc_sheet(wb)

    wb.save(output_path)
    print(f"Created workbook: {output_path}")
    print("Extracted points:")
    for p, flag in zip(points, use_flags, strict=True):
        print(
            f"  conc={p.concentration_mg_ml:.4g} mg/mL, abs={p.absorbance:.6f}, use_for_fit={flag}"
        )


if __name__ == "__main__":
    main()
